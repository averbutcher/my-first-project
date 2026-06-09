import re
import json
from datetime import datetime, timedelta, time as time_type
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

CONFIG_PATH = Path(__file__).parent / "config.json"

DEFAULT_CONFIG = {
    "excel_columns": {
        "date": "B",
        "start_time": "C",
        "end_time": "D",
        "worker_name": "F"
    },
    "excel_has_header": True,
    "rules": {
        "gap_threshold_minutes": 30,
        "default_start_time": "10:00"
    },
    "aliases": {}
}


def load_config():
    if not CONFIG_PATH.exists():
        save_config(DEFAULT_CONFIG)
        return dict(DEFAULT_CONFIG)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def col_to_idx(letter: str) -> int:
    return ord(letter.strip().upper()) - ord("A")


def _is_na(val) -> bool:
    try:
        return bool(pd.isna(val))
    except (TypeError, ValueError):
        return False


def parse_time(val) -> time_type | None:
    if val is None or _is_na(val):
        return None
    if isinstance(val, time_type):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(val.strip(), fmt).time()
            except ValueError:
                pass
        return None
    if isinstance(val, (int, float)):
        # Excel stores times as fractions of a day
        if 0 <= val <= 1:
            total_sec = round(val * 86400)
            h, rem = divmod(total_sec, 3600)
            m = rem // 60
            return time_type(h % 24, m)
    return None


def time_to_hours(t: time_type) -> float:
    return t.hour + t.minute / 60 + t.second / 3600


def hours_between(start: time_type, end: time_type) -> float:
    diff = time_to_hours(end) - time_to_hours(start)
    if diff < 0:
        diff += 24  # overnight shift
    return diff


def add_hours(t: time_type, h: float) -> time_type:
    dt = datetime.combine(datetime.today(), t) + timedelta(hours=h)
    return dt.time()


# ── Message parsing ────────────────────────────────────────────────────────────

def parse_message(text: str) -> list[dict]:
    entries = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split(",", 2)
        if len(parts) < 3:
            continue

        worker_name = parts[0].strip()
        workplace   = parts[1].strip()
        rest        = parts[2].strip()

        # Detect direction: number-then-letter (8ש) or letter-then-number (ש8)
        if re.search(r"\d+ש", rest):
            hours_m = re.search(r"(\d+(?:\.\d+)?)ש", rest)
            sales_m = re.search(r"(\d+(?:\.\d+)?)מ", rest)
            hours = float(hours_m.group(1)) if hours_m else None
            sales = int(float(sales_m.group(1))) if sales_m else ""
        else:
            hours_m = re.search(r"ש(\d+(?:\.\d+)?)", rest)
            sales_m = re.search(r"מ(\d+(?:\.\d+)?)", rest)
            hours = float(hours_m.group(1)) if hours_m else None
            sales = int(float(sales_m.group(1))) if sales_m else ""

        if worker_name:
            entries.append({
                "worker_name": worker_name,
                "workplace":   workplace,
                "hours":       hours,
                "sales":       sales,
            })
    return entries


# ── Excel parsing ──────────────────────────────────────────────────────────────

def _is_ignored(name: str, ignored_names: list) -> bool:
    """
    Returns True if 'name' matches any entry in ignored_names.
    Handles word reordering and partial names (e.g. first name only).
    """
    name_words = set(name.strip().split())
    for ignored in ignored_names:
        ig_words = set(ignored.split())
        if not ig_words:
            continue
        # All words of the shorter side appear in the longer side
        if ig_words.issubset(name_words) or name_words.issubset(ig_words):
            return True
    return False


def parse_excel(file_path: str, cfg: dict) -> list[dict]:
    cols          = cfg["excel_columns"]
    has_header    = cfg.get("excel_has_header", True)
    ignored_names = [n.strip() for n in cfg.get("ignored_names", []) if n.strip()]

    date_idx  = col_to_idx(cols["date"])
    start_idx = col_to_idx(cols["start_time"])
    end_idx   = col_to_idx(cols["end_time"])
    name_idx  = col_to_idx(cols["worker_name"])

    skip = 1 if has_header else 0
    df   = pd.read_excel(file_path, header=None, skiprows=skip)

    entries = []
    for _, row in df.iterrows():
        try:
            name = str(row.iloc[name_idx]).strip()
            if not name or name.lower() == "nan":
                continue
            if _is_ignored(name, ignored_names):
                continue

            date_val = row.iloc[date_idx]
            start_t  = parse_time(row.iloc[start_idx])
            end_t    = parse_time(row.iloc[end_idx])

            # Skip rows with no shift times
            if start_t is None or end_t is None:
                continue

            hours = hours_between(start_t, end_t)

            entries.append({
                "worker_name": name,
                "date":        date_val,
                "start_time":  start_t,
                "end_time":    end_t,
                "hours":       hours,
            })
        except Exception:
            continue
    return entries


# ── Name matching ──────────────────────────────────────────────────────────────

def canonical_key(name: str, aliases: dict) -> str:
    """Map a name to its canonical group key using aliases."""
    for k, v in aliases.items():
        if name == k or name == v:
            return k
    return name


def _first_name(name: str) -> str:
    return name.split()[0] if name.strip() else name


def _build_word_lookup(excel_by_key: dict, word_index: int) -> dict[str, str | None]:
    """
    Build {word: canonical_key} where 'word' is the word at word_index in the
    Excel name. Keys that map to more than one worker are set to None (ambiguous).
    """
    lookup: dict[str, str | None] = {}
    for key in excel_by_key:
        words = key.split()
        if len(words) <= word_index:
            continue
        word = words[word_index]
        if word in lookup:
            lookup[word] = None  # ambiguous
        else:
            lookup[word] = key
    return lookup


# ── Comparison ─────────────────────────────────────────────────────────────────

def compare(msg_entries: list, excel_entries: list, cfg: dict) -> list[dict]:
    aliases       = cfg.get("aliases", {})
    threshold_min = cfg["rules"]["gap_threshold_minutes"]
    default_start = datetime.strptime(cfg["rules"]["default_start_time"], "%H:%M").time()

    # Infer "the date" for this run from Excel
    excel_date = next(
        (e["date"] for e in excel_entries if not _is_na(e["date"])),
        None
    )

    excel_by_key: dict[str, list] = {}
    for e in excel_entries:
        k = canonical_key(e["worker_name"], aliases)
        excel_by_key.setdefault(k, []).append(e)

    # Name fallback lookups (word 0 = first word, word 1 = second word of Excel name)
    first_word_lookup  = _build_word_lookup(excel_by_key, 0)  # Excel first name first
    second_word_lookup = _build_word_lookup(excel_by_key, 1)  # Excel last name first

    def resolve_msg_key(msg_name: str) -> str:
        """
        Matching priority:
        1. Alias
        2. Exact name
        3. Message first word == Excel first word  (e.g. "ישראל" → "ישראל ישראלי")
        4. Message first word == Excel second word (e.g. "ישראל" → "ישראלי ישראל")
        """
        k = canonical_key(msg_name, aliases)
        if k in excel_by_key:
            return k
        fn = _first_name(msg_name)
        matched = first_word_lookup.get(fn)
        if matched is not None:
            return matched
        matched = second_word_lookup.get(fn)
        if matched is not None:
            return matched
        return k

    msg_by_key: dict[str, list] = {}
    for e in msg_entries:
        k = resolve_msg_key(e["worker_name"])
        msg_by_key.setdefault(k, []).append(e)

    all_keys = sorted(set(excel_by_key) | set(msg_by_key))
    output   = []

    for key in all_keys:
        ex_rows = excel_by_key.get(key, [])
        ms_rows = msg_by_key.get(key, [])
        count   = max(len(ex_rows), len(ms_rows))

        for i in range(count):
            er = ex_rows[i] if i < len(ex_rows) else None
            mr = ms_rows[i] if i < len(ms_rows) else None

            if er and mr:
                eh, mh = er["hours"], mr["hours"]

                if mh is None:
                    # Message line had no hours — use Excel times as-is
                    output.append({
                        "date":        er["date"],
                        "worker_name": er["worker_name"],
                        "workplace":   mr["workplace"],
                        "start_time":  er["start_time"],
                        "end_time":    er["end_time"],
                        "sales":       mr["sales"],
                        "notes":       "שעות חסרות בהודעה",
                        "status":      "gap",
                    })
                    continue

                diff_min = abs(eh - mh) * 60

                if diff_min <= threshold_min:
                    # Rule 1 — times match
                    output.append({
                        "date":        er["date"],
                        "worker_name": er["worker_name"],
                        "workplace":   mr["workplace"],
                        "start_time":  er["start_time"],
                        "end_time":    er["end_time"],
                        "sales":       mr["sales"],
                        "notes":       "הכל תקין",
                        "status":      "ok",
                    })
                else:
                    # Rule 2 — gap: use Excel start, message hours for end
                    output.append({
                        "date":        er["date"],
                        "worker_name": er["worker_name"],
                        "workplace":   mr["workplace"],
                        "start_time":  er["start_time"],
                        "end_time":    add_hours(er["start_time"], mh),
                        "sales":       mr["sales"],
                        "notes":       f"פער של {int(diff_min)} דקות",
                        "status":      "gap",
                    })

            elif er:
                # Rule 3 — missing from message
                output.append({
                    "date":        er["date"],
                    "worker_name": er["worker_name"],
                    "workplace":   "",
                    "start_time":  er["start_time"],
                    "end_time":    er["end_time"],
                    "sales":       "",
                    "notes":       "חסר בהודעה",
                    "status":      "missing_msg",
                })

            else:
                # Rule 4 — missing from Excel
                mh = mr["hours"] or 0
                output.append({
                    "date":        excel_date or "",
                    "worker_name": mr["worker_name"],
                    "workplace":   mr["workplace"],
                    "start_time":  default_start,
                    "end_time":    add_hours(default_start, mh),
                    "sales":       mr["sales"],
                    "notes":       "חסר באקסל",
                    "status":      "missing_excel",
                })

    return output


# ── Export ─────────────────────────────────────────────────────────────────────

def _fmt_time(t) -> str:
    if isinstance(t, time_type):
        return t.strftime("%H:%M")
    return str(t) if t else ""


def _fmt_date(d) -> str:
    if isinstance(d, datetime):
        return d.strftime("%d/%m/%Y")
    if hasattr(d, "strftime"):
        return d.strftime("%d/%m/%Y")
    return str(d) if (d and not _is_na(d)) else ""


STATUS_COLORS = {
    "ok":            "E2EFDA",  # green
    "gap":           "FFF2CC",  # yellow
    "missing_msg":   "FCE4D6",  # orange
    "missing_excel": "FFDADA",  # red
}

HEADERS = ["תאריך", "שם עובד", "מקום עבודה", "שעת התחלה", "שעת סיום", "מכירות", "הערות"]


def export_to_excel(rows: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "השוואה"
    ws.sheet_view.rightToLeft = True

    ws.append(HEADERS)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            _fmt_date(row["date"]),
            row["worker_name"],
            row["workplace"],
            _fmt_time(row["start_time"]),
            _fmt_time(row["end_time"]),
            row["sales"],
            row["notes"],
        ])
        color = STATUS_COLORS.get(row["status"], "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(width + 4, 14)

    wb.save(output_path)
